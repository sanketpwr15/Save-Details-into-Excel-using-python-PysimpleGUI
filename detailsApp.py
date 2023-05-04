import PySimpleGUI as sg
from datetime import datetime
import pathlib
from openpyxl import Workbook
import openpyxl
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
import os

#create excel file and give them headers and save it to location 
file = pathlib.Path("backend_Data.xlsx")
if file.exists ():

    pass

else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]= "FIRST NAME"
    sheet["B1"]= "LAST NAME"
    sheet["C1"]= "NUMBER"
    sheet["D1"]= "BIRTH DATE"
    sheet["E1"]= "ADDRESS"
file.save("backend_Data.xlsx")


# windoews theam 
sg.theme('DarkAmber')

# Set the icon for the layout
icon_path = os.path.abspath("person1.ico")
sg.set_global_icon(icon_path)

# All the stuff inside your window.
layout = [  [sg.Text('Details', enable_events=True,
                        key='-TEXT-', font=('Arial bold', 15),
                        expand_x=True, justification='center') ],
                        #First name:
                        [sg.T("")], 
                        [sg.Text("First Name:",size =(10, 1)), sg.Input(size=(30, 1),key="-fname-" )],
                        #last name:
                        [sg.T("")],
                        [sg.Text("Last Name:",size =(10, 1)), sg.Input(size=(30, 1),key="-lname-")],
                        #phone Number:
                        [sg.T("")],
                        [sg.Text("Number:",size =(10, 1)), sg.Input(size=(30, 1),key="-pnum-")],
                        #birth date:
                        [sg.T("")],
                        [sg.Text("Birth Date:",size =(10, 1)),sg.CalendarButton("Date", 
                        close_when_date_chosen=True,  target='-date-',
                        location=(0,0), no_titlebar=False ),(sg.Input(key='-date-', size=(22,1)))],
                        #Address
                        [sg.T("")],
                        [sg.Text("Address",size=(10, 1)),sg.Multiline(key='-address-', size=(30,3))],
                        #buttons
                        [sg.T("")],
                        [sg.Button('Save',button_color = 'Green'),
                        sg.Button('Clear',button_color = 'red'),
                        sg.Button('Exit'),
                        sg.Button('Show Details')]
         ]

# Create the Window
window = sg.Window('Personal Details', layout,size=(370,450))

# Event Loop to process "events" and button functions
while True:
    event, values = window.read()
    # if user closes window or clicks cancel
    if event == sg.WIN_CLOSED or event == 'Exit':
        sg.popup('Do You Want to Close this Window')
        break
        
        
    
    #Save button clicked
    if event == 'Save':
        file = openpyxl.load_workbook("backend_Data.xlsx")
        sheet=file.active
        sheet.cell(column=1, row=sheet.max_row+1,value=values['-fname-'])
        sheet.cell(column=2, row=sheet.max_row,value=values['-lname-'])
        sheet.cell(column=3, row=sheet.max_row,value=values["-pnum-"])
        sheet.cell(column=4, row=sheet.max_row,value=values["-date-"])
        sheet.cell(column=5, row=sheet.max_row,value=values["-address-"])
        file.save("backend_Data.xlsx")
        sg.popup('Details Save Successfully')
    
    #Clear button clicked    
    if event == 'Clear':
        window['-fname-'].update('')
        window['-lname-'].update('')
        window['-pnum-'].update('')
        window['-date-'].update('')
        window['-address-'].update('')
        
    
    #Show Details button clicked
    if event == 'Show Details':
        df = pd.read_excel(pathlib.Path("backend_Data.xlsx"))
        data_str = df.to_string(index=False)

        # Create a list of column names
        col_names = df.columns.tolist()

        # Create a list of row data
        row_data = df.values.tolist()
    #Details Layout 
        layout = [[sg.Table(values=row_data, headings=col_names, max_col_width=25, auto_size_columns=True,
                justification='center', num_rows=min(25, len(row_data)))],
                #Close and Search Button
                [sg.Button('Back'),sg.Button('Export as PDF'),[sg.Text('Search:'), sg.InputText(key='-SEARCH-'), sg.Button('Search')],
                [sg.Table(values=df.values.tolist(), headings=df.columns.tolist(), key='-TABLE-')]]]

        Details_layout = sg.Window('Details', layout, grab_anywhere=False)
        while True:
        # Read events from the sublayout window
            sub_event, sub_values = Details_layout.read()

        # Handle events
            if sub_event == sg.WIN_CLOSED:
                break
            elif sub_event == 'Back':
            # Close the sublayout window
                Details_layout.close()
                break

        # Search Button Function
            if sub_event == 'Search':
            # Get the search term from the search bar
                search_term = sub_values['-SEARCH-']

            # Filter the DataFrame to match the search term
                filtered_df = df[df.apply(lambda row: search_term.lower() in row.values.astype(str).tolist(), axis=1)]

            # Update the table with the filtered results
                Details_layout['-TABLE-'].update(values=filtered_df.values.tolist())
        

            if sub_event == 'Export as PDF':
        # Create the PDF file
                filename = sg.popup_get_file('Save As', save_as=True, file_types=(('PDF Files', '*.pdf'),))
                if filename:
            # Create a PDF document
                    doc = SimpleDocTemplate(filename, pagesize=letter)
                    elements = []

            # Create a table style
                    style = TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 14),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 12),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ])

                    # Create the table
                    data = [df.columns.tolist()] + filtered_df.values.tolist()
                    table = Table(data)
                    table.setStyle(style)

                    # Add the table to the document
                    elements.append(table)

                    # Build the PDF document and save the file
                    doc.build(elements)
                    sg.popup('PDF exported successfully!')
        
        Details_layout.close()
window.close()